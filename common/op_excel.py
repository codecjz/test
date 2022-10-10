from openpyxl import load_workbook
from common.util_cl import Util

class OperateExcel:
    #获取文档
    def __init__(self,fn=None):
        if not fn:
            self.fn = Util.get_conf()['testCaseDir']
        else:
            self.fn = fn
        self.wb = load_workbook(self.fn)
    #获取测试用例表格对象
    def get_data(self):
        ws = self.wb.active
        return ws
    #获取数据的总行数
    def get_lines(self):
        return len(list(self.get_data().rows))#表格的每一行转换为列表的每一项

    #获取每个单元格内容
    def get_value(self,row,col):
        return self.get_data().cell(row=row,column=col).value
    

if __name__ == '__main__':
    oe = OperateExcel()
    print(oe.get_value(2, 2))